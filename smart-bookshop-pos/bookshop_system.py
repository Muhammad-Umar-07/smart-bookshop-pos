"""
Smart Book Shop Management & Billing System
Professional POS Software for Book Sellers
Using customtkinter for modern, clean UI
"""

import customtkinter as ctk
from tkinter import messagebox, ttk
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

# Set appearance and color theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class BookShopSystem:
    """Main application class for Book Shop Management System"""
    
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("Smart Book Shop Management & Billing System")
        self.root.geometry("1200x700")
        
        # Initialize data storage
        self.setup_directories()
        self.load_credentials()
        self.load_inventory()
        
        # Current user state
        self.logged_in = False
        self.current_cart = []
        
        # Show login screen
        self.show_login_screen()
        
    def setup_directories(self):
        """Create necessary folders for the system"""
        folders = ['Inventory', 'Sales_Records', 'Application_Files']
        for folder in folders:
            os.makedirs(folder, exist_ok=True)
    
    def load_credentials(self):
        """Load or create staff credentials"""
        self.credentials_file = 'Application_Files/credentials.json'
        if os.path.exists(self.credentials_file):
            with open(self.credentials_file, 'r') as f:
                self.credentials = json.load(f)
        else:
            # Default password
            self.credentials = {'password': 'admin123'}
            self.save_credentials()
    
    def save_credentials(self):
        """Save staff credentials"""
        with open(self.credentials_file, 'w') as f:
            json.dump(self.credentials, f, indent=4)
    
    def load_inventory(self):
        """Load book inventory from file"""
        self.inventory_file = 'Inventory/books.json'
        if os.path.exists(self.inventory_file):
            with open(self.inventory_file, 'r') as f:
                self.books = json.load(f)
        else:
            self.books = []
            self.save_inventory()
    
    def save_inventory(self):
        """Save book inventory to file"""
        with open(self.inventory_file, 'w') as f:
            json.dump(self.books, f, indent=4)
    
    def clear_screen(self):
        """Clear all widgets from the window"""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    # ============ LOGIN SYSTEM ============
    
    def show_login_screen(self):
        """Display staff login screen"""
        self.clear_screen()
        
        # Main container
        container = ctk.CTkFrame(self.root)
        container.pack(expand=True)
        
        # Title
        title = ctk.CTkLabel(
            container,
            text="📚 Smart Book Shop System",
            font=("Arial", 32, "bold")
        )
        title.pack(pady=30)
        
        subtitle = ctk.CTkLabel(
            container,
            text="Staff Login",
            font=("Arial", 20)
        )
        subtitle.pack(pady=10)
        
        # Password frame
        pass_frame = ctk.CTkFrame(container)
        pass_frame.pack(pady=30)
        
        ctk.CTkLabel(
            pass_frame,
            text="Password:",
            font=("Arial", 16)
        ).pack(pady=10)
        
        self.password_entry = ctk.CTkEntry(
            pass_frame,
            width=300,
            height=40,
            font=("Arial", 14),
            show="●"
        )
        self.password_entry.pack(pady=10)
        self.password_entry.bind('<Return>', lambda e: self.login())
        
        # Login button
        ctk.CTkButton(
            pass_frame,
            text="Login",
            width=200,
            height=45,
            font=("Arial", 16, "bold"),
            command=self.login
        ).pack(pady=20)
        
        self.password_entry.focus()
    
    def login(self):
        """Validate login credentials"""
        password = self.password_entry.get()
        
        if password == self.credentials['password']:
            self.logged_in = True
            self.show_main_menu()
        else:
            messagebox.showerror(
                "Login Failed",
                "Incorrect password. Please try again."
            )
            self.password_entry.delete(0, 'end')
    
    # ============ MAIN MENU ============
    
    def show_main_menu(self):
        """Display main menu with professional options"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=80)
        header.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(
            header,
            text="📚 Book Shop Management System",
            font=("Arial", 28, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="Staff Logout",
            width=150,
            height=40,
            font=("Arial", 14),
            fg_color="#dc3545",
            hover_color="#c82333",
            command=self.logout
        ).pack(side="right", padx=20)
        
        # Menu buttons container
        menu_frame = ctk.CTkFrame(self.root)
        menu_frame.pack(expand=True, fill="both", padx=100, pady=50)
        
        buttons = [
            ("🛒 New Sale", self.show_new_sale, "#28a745"),
            ("📦 Inventory Management", self.show_inventory_menu, "#007bff"),
            ("📊 Sales Reports", self.show_sales_reports, "#17a2b8"),
            ("🔐 Change Password", self.show_change_password, "#ffc107"),
            ("❌ Exit System", self.exit_system, "#6c757d")
        ]
        
        for i, (text, command, color) in enumerate(buttons):
            btn = ctk.CTkButton(
                menu_frame,
                text=text,
                width=400,
                height=70,
                font=("Arial", 18, "bold"),
                fg_color=color,
                command=command
            )
            btn.pack(pady=15)
    
    def logout(self):
        """Logout with password confirmation"""
        self.logged_in = False
        self.current_cart = []
        messagebox.showinfo("Logout", "You have been logged out successfully.")
        self.show_login_screen()
    
    def exit_system(self):
        """Exit the application"""
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            self.root.quit()
    
    # ============ INVENTORY MANAGEMENT ============
    
    def show_inventory_menu(self):
        """Display inventory management options"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="📦 Inventory Management",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_main_menu
        ).pack(side="right", padx=20)
        
        # Menu options
        menu_frame = ctk.CTkFrame(self.root)
        menu_frame.pack(expand=True, fill="both", padx=80, pady=40)
        
        buttons = [
            ("➕ Add New Book", self.show_add_book),
            ("📝 Edit Book", self.show_edit_book),
            ("🗑️ Delete Book", self.show_delete_book),
            ("📚 View All Books", self.show_view_books)
        ]
        
        for text, command in buttons:
            ctk.CTkButton(
                menu_frame,
                text=text,
                width=400,
                height=60,
                font=("Arial", 16, "bold"),
                command=command
            ).pack(pady=12)
    
    def show_add_book(self):
        """Display form to add a new book"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="➕ Add New Book",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_inventory_menu
        ).pack(side="right", padx=20)
        
        # Form container
        form_frame = ctk.CTkFrame(self.root)
        form_frame.pack(expand=True, padx=100, pady=20)
        
        # Form fields
        fields = [
            ("Book Title:", "title"),
            ("SKU / Serial Number:", "sku"),
            ("Category/Class (9, 10, 11, 12):", "category"),
            ("Unit Price (Rs):", "price")
        ]
        
        self.add_book_entries = {}
        
        for label, key in fields:
            row = ctk.CTkFrame(form_frame)
            row.pack(pady=15, fill="x", padx=50)
            
            ctk.CTkLabel(
                row,
                text=label,
                font=("Arial", 16),
                width=250,
                anchor="w"
            ).pack(side="left", padx=10)
            
            entry = ctk.CTkEntry(
                row,
                width=350,
                height=40,
                font=("Arial", 14)
            )
            entry.pack(side="left", padx=10)
            self.add_book_entries[key] = entry
        
        # Add button
        ctk.CTkButton(
            form_frame,
            text="Add Book",
            width=250,
            height=50,
            font=("Arial", 16, "bold"),
            fg_color="#28a745",
            command=self.add_book
        ).pack(pady=30)
    
    def add_book(self):
        """Add a new book to inventory"""
        try:
            # Get values
            title = self.add_book_entries['title'].get().strip()
            sku = self.add_book_entries['sku'].get().strip()
            category = self.add_book_entries['category'].get().strip()
            price = self.add_book_entries['price'].get().strip()
            
            # Validation
            if not all([title, sku, category, price]):
                messagebox.showerror("Error", "Please fill in all fields.")
                return
            
            # Check if SKU already exists
            if any(book['sku'] == sku for book in self.books):
                messagebox.showerror("Error", f"SKU '{sku}' already exists!")
                return
            
            # Validate category
            if category not in ['9', '10', '11', '12']:
                messagebox.showerror("Error", "Category must be 9, 10, 11, or 12.")
                return
            
            # Validate price
            try:
                price_value = float(price)
                if price_value <= 0:
                    raise ValueError()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid price.")
                return
            
            # Add book
            book = {
                'title': title,
                'sku': sku,
                'category': category,
                'price': price_value
            }
            
            self.books.append(book)
            self.save_inventory()
            
            messagebox.showinfo("Success", f"Book '{title}' added successfully!")
            
            # Clear form
            for entry in self.add_book_entries.values():
                entry.delete(0, 'end')
            
            self.add_book_entries['title'].focus()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add book: {str(e)}")
    
    def show_view_books(self):
        """Display all books in a table"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="📚 Book Inventory",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_inventory_menu
        ).pack(side="right", padx=20)
        
        # Search and filter frame
        filter_frame = ctk.CTkFrame(self.root)
        filter_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            filter_frame,
            text="Filter by Class:",
            font=("Arial", 14)
        ).pack(side="left", padx=10)
        
        self.filter_var = ctk.StringVar(value="All")
        filter_menu = ctk.CTkOptionMenu(
            filter_frame,
            values=["All", "9", "10", "11", "12"],
            variable=self.filter_var,
            command=lambda x: self.update_book_table()
        )
        filter_menu.pack(side="left", padx=10)
        
        ctk.CTkLabel(
            filter_frame,
            text="Search:",
            font=("Arial", 14)
        ).pack(side="left", padx=(30, 10))
        
        self.search_entry = ctk.CTkEntry(filter_frame, width=300)
        self.search_entry.pack(side="left", padx=10)
        self.search_entry.bind('<KeyRelease>', lambda e: self.update_book_table())
        
        # Table frame
        table_frame = ctk.CTkFrame(self.root)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Create Treeview
        columns = ('SKU', 'Title', 'Category', 'Price')
        self.books_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=20
        )
        
        # Column headings
        self.books_tree.heading('SKU', text='SKU / Serial')
        self.books_tree.heading('Title', text='Book Title')
        self.books_tree.heading('Category', text='Class')
        self.books_tree.heading('Price', text='Unit Price (Rs)')
        
        # Column widths
        self.books_tree.column('SKU', width=150)
        self.books_tree.column('Title', width=400)
        self.books_tree.column('Category', width=100)
        self.books_tree.column('Price', width=150)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.books_tree.yview)
        self.books_tree.configure(yscrollcommand=scrollbar.set)
        
        self.books_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Info label
        self.book_count_label = ctk.CTkLabel(
            self.root,
            text="",
            font=("Arial", 14)
        )
        self.book_count_label.pack(pady=10)
        
        self.update_book_table()
    
    def update_book_table(self):
        """Update the book table with filters"""
        # Clear existing items
        for item in self.books_tree.get_children():
            self.books_tree.delete(item)
        
        # Get filter values
        category_filter = self.filter_var.get()
        search_term = self.search_entry.get().lower()
        
        # Filter books
        filtered_books = self.books
        
        if category_filter != "All":
            filtered_books = [b for b in filtered_books if b['category'] == category_filter]
        
        if search_term:
            filtered_books = [
                b for b in filtered_books
                if search_term in b['title'].lower() or search_term in b['sku'].lower()
            ]
        
        # Populate table
        for book in filtered_books:
            self.books_tree.insert('', 'end', values=(
                book['sku'],
                book['title'],
                f"Class {book['category']}",
                f"Rs {book['price']:.2f}"
            ))
        
        # Update count
        self.book_count_label.configure(
            text=f"Showing {len(filtered_books)} of {len(self.books)} books"
        )
    
    def show_edit_book(self):
        """Display interface to select and edit a book"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="📝 Edit Book",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_inventory_menu
        ).pack(side="right", padx=20)
        
        if not self.books:
            ctk.CTkLabel(
                self.root,
                text="No books in inventory. Please add books first.",
                font=("Arial", 18)
            ).pack(expand=True)
            return
        
        # Selection frame
        select_frame = ctk.CTkFrame(self.root)
        select_frame.pack(fill="both", expand=True, padx=50, pady=20)
        
        ctk.CTkLabel(
            select_frame,
            text="Select a book to edit:",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        # Book list
        list_frame = ctk.CTkScrollableFrame(select_frame, height=400)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        for book in self.books:
            btn_text = f"{book['title']} | SKU: {book['sku']} | Class {book['category']} | Rs {book['price']:.2f}"
            ctk.CTkButton(
                list_frame,
                text=btn_text,
                width=700,
                height=50,
                font=("Arial", 14),
                anchor="w",
                command=lambda b=book: self.edit_book_form(b)
            ).pack(pady=5)
    
    def edit_book_form(self, book):
        """Display edit form for selected book"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text=f"📝 Editing: {book['title']}",
            font=("Arial", 20, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_edit_book
        ).pack(side="right", padx=20)
        
        # Form
        form_frame = ctk.CTkFrame(self.root)
        form_frame.pack(expand=True, padx=100, pady=20)
        
        fields = [
            ("Book Title:", "title", book['title']),
            ("SKU / Serial Number:", "sku", book['sku']),
            ("Category/Class:", "category", book['category']),
            ("Unit Price (Rs):", "price", str(book['price']))
        ]
        
        self.edit_entries = {}
        
        for label, key, value in fields:
            row = ctk.CTkFrame(form_frame)
            row.pack(pady=15, fill="x", padx=50)
            
            ctk.CTkLabel(
                row,
                text=label,
                font=("Arial", 16),
                width=250,
                anchor="w"
            ).pack(side="left", padx=10)
            
            entry = ctk.CTkEntry(
                row,
                width=350,
                height=40,
                font=("Arial", 14)
            )
            entry.insert(0, value)
            entry.pack(side="left", padx=10)
            self.edit_entries[key] = entry
        
        # Save button
        ctk.CTkButton(
            form_frame,
            text="Save Changes",
            width=250,
            height=50,
            font=("Arial", 16, "bold"),
            fg_color="#28a745",
            command=lambda: self.update_book(book['sku'])
        ).pack(pady=30)
    
    def update_book(self, original_sku):
        """Update book information"""
        try:
            # Get new values
            title = self.edit_entries['title'].get().strip()
            new_sku = self.edit_entries['sku'].get().strip()
            category = self.edit_entries['category'].get().strip()
            price = self.edit_entries['price'].get().strip()
            
            # Validation
            if not all([title, new_sku, category, price]):
                messagebox.showerror("Error", "Please fill in all fields.")
                return
            
            # Check SKU conflict
            if new_sku != original_sku:
                if any(book['sku'] == new_sku for book in self.books):
                    messagebox.showerror("Error", f"SKU '{new_sku}' already exists!")
                    return
            
            if category not in ['9', '10', '11', '12']:
                messagebox.showerror("Error", "Category must be 9, 10, 11, or 12.")
                return
            
            try:
                price_value = float(price)
                if price_value <= 0:
                    raise ValueError()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid price.")
                return
            
            # Update book
            for book in self.books:
                if book['sku'] == original_sku:
                    book['title'] = title
                    book['sku'] = new_sku
                    book['category'] = category
                    book['price'] = price_value
                    break
            
            self.save_inventory()
            messagebox.showinfo("Success", "Book updated successfully!")
            self.show_inventory_menu()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update book: {str(e)}")
    
    def show_delete_book(self):
        """Display interface to select and delete a book"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="🗑️ Delete Book",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_inventory_menu
        ).pack(side="right", padx=20)
        
        if not self.books:
            ctk.CTkLabel(
                self.root,
                text="No books in inventory.",
                font=("Arial", 18)
            ).pack(expand=True)
            return
        
        # Selection frame
        select_frame = ctk.CTkFrame(self.root)
        select_frame.pack(fill="both", expand=True, padx=50, pady=20)
        
        ctk.CTkLabel(
            select_frame,
            text="Select a book to delete:",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        # Book list
        list_frame = ctk.CTkScrollableFrame(select_frame, height=500)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        for book in self.books:
            btn_text = f"{book['title']} | SKU: {book['sku']} | Class {book['category']} | Rs {book['price']:.2f}"
            
            btn_frame = ctk.CTkFrame(list_frame)
            btn_frame.pack(fill="x", pady=5)
            
            ctk.CTkLabel(
                btn_frame,
                text=btn_text,
                font=("Arial", 14),
                anchor="w"
            ).pack(side="left", fill="x", expand=True, padx=10)
            
            ctk.CTkButton(
                btn_frame,
                text="Delete",
                width=100,
                height=35,
                fg_color="#dc3545",
                command=lambda b=book: self.delete_book(b)
            ).pack(side="right", padx=10)
    
    def delete_book(self, book):
        """Delete a book after confirmation"""
        if messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete:\n\n{book['title']} (SKU: {book['sku']})?"):
            
            self.books = [b for b in self.books if b['sku'] != book['sku']]
            self.save_inventory()
            messagebox.showinfo("Success", "Book deleted successfully!")
            self.show_delete_book()
    
    # ============ SALES / BILLING SYSTEM ============
    
    def show_new_sale(self):
        """Display new sale / billing interface"""
        self.clear_screen()
        self.current_cart = []
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="🛒 New Sale",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_main_menu
        ).pack(side="right", padx=20)
        
        if not self.books:
            ctk.CTkLabel(
                self.root,
                text="No books available. Please add books to inventory first.",
                font=("Arial", 18)
            ).pack(expand=True)
            return
        
        # Main container
        main_container = ctk.CTkFrame(self.root)
        main_container.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Left side - Book selection
        left_frame = ctk.CTkFrame(main_container)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        ctk.CTkLabel(
            left_frame,
            text="Available Books",
            font=("Arial", 18, "bold")
        ).pack(pady=10)
        
        # Filter
        filter_frame = ctk.CTkFrame(left_frame)
        filter_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            filter_frame,
            text="Filter by Class:",
            font=("Arial", 14)
        ).pack(side="left", padx=5)
        
        self.sale_filter_var = ctk.StringVar(value="All")
        ctk.CTkOptionMenu(
            filter_frame,
            values=["All", "9", "10", "11", "12"],
            variable=self.sale_filter_var,
            command=lambda x: self.update_sale_books()
        ).pack(side="left", padx=5)
        
        # Book list
        self.sale_books_frame = ctk.CTkScrollableFrame(left_frame, height=450)
        self.sale_books_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Right side - Cart
        right_frame = ctk.CTkFrame(main_container, width=400)
        right_frame.pack(side="right", fill="both", padx=(10, 0))
        
        ctk.CTkLabel(
            right_frame,
            text="Shopping Cart",
            font=("Arial", 18, "bold")
        ).pack(pady=10)
        
        # Cart display
        self.cart_frame = ctk.CTkScrollableFrame(right_frame, height=350)
        self.cart_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Total
        self.total_label = ctk.CTkLabel(
            right_frame,
            text="Total: Rs 0.00",
            font=("Arial", 20, "bold")
        )
        self.total_label.pack(pady=15)
        
        # Buttons
        btn_frame = ctk.CTkFrame(right_frame)
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(
            btn_frame,
            text="Clear Cart",
            height=40,
            fg_color="#dc3545",
            command=self.clear_cart
        ).pack(fill="x", pady=5)
        
        ctk.CTkButton(
            btn_frame,
            text="Generate Invoice",
            height=50,
            font=("Arial", 16, "bold"),
            fg_color="#28a745",
            command=self.generate_invoice
        ).pack(fill="x", pady=5)
        
        self.update_sale_books()
        self.update_cart_display()
    
    def update_sale_books(self):
        """Update the available books list for sale"""
        # Clear existing
        for widget in self.sale_books_frame.winfo_children():
            widget.destroy()
        
        # Filter books
        category_filter = self.sale_filter_var.get()
        filtered_books = self.books
        
        if category_filter != "All":
            filtered_books = [b for b in self.books if b['category'] == category_filter]
        
        # Display books
        for book in filtered_books:
            btn_frame = ctk.CTkFrame(self.sale_books_frame)
            btn_frame.pack(fill="x", pady=5)
            
            text = f"{book['title']}\nClass {book['category']} | SKU: {book['sku']}\nRs {book['price']:.2f}"
            
            ctk.CTkButton(
                btn_frame,
                text=text,
                height=70,
                anchor="w",
                command=lambda b=book: self.add_to_cart(b)
            ).pack(fill="x", padx=5)
    
    def add_to_cart(self, book):
        """Add a book to the shopping cart"""
        self.current_cart.append(book.copy())
        self.update_cart_display()
        messagebox.showinfo("Added", f"Added '{book['title']}' to cart!")
    
    def update_cart_display(self):
        """Update the cart display"""
        # Clear existing
        for widget in self.cart_frame.winfo_children():
            widget.destroy()
        
        if not self.current_cart:
            ctk.CTkLabel(
                self.cart_frame,
                text="Cart is empty",
                font=("Arial", 14),
                text_color="gray"
            ).pack(pady=20)
        else:
            for i, book in enumerate(self.current_cart):
                item_frame = ctk.CTkFrame(self.cart_frame)
                item_frame.pack(fill="x", pady=5)
                
                info = ctk.CTkLabel(
                    item_frame,
                    text=f"{book['title']}\nClass {book['category']} | Rs {book['price']:.2f}",
                    font=("Arial", 12),
                    anchor="w"
                )
                info.pack(side="left", fill="x", expand=True, padx=5)
                
                ctk.CTkButton(
                    item_frame,
                    text="×",
                    width=30,
                    height=30,
                    fg_color="#dc3545",
                    command=lambda idx=i: self.remove_from_cart(idx)
                ).pack(side="right", padx=5)
        
        # Update total
        total = sum(book['price'] for book in self.current_cart)
        self.total_label.configure(text=f"Total: Rs {total:.2f}")
    
    def remove_from_cart(self, index):
        """Remove an item from cart"""
        self.current_cart.pop(index)
        self.update_cart_display()
    
    def clear_cart(self):
        """Clear all items from cart"""
        if self.current_cart:
            if messagebox.askyesno("Clear Cart", "Remove all items from cart?"):
                self.current_cart = []
                self.update_cart_display()
    
    def generate_invoice(self):
        """Generate invoice and save to Excel"""
        if not self.current_cart:
            messagebox.showwarning("Empty Cart", "Please add items to cart first!")
            return
        
        try:
            # Calculate totals
            total_amount = sum(book['price'] for book in self.current_cart)
            total_books = len(self.current_cart)
            
            # Get current date/time
            now = datetime.now()
            date_str = now.strftime("%d-%m-%Y")
            time_str = now.strftime("%I:%M %p")
            
            # Save to Excel
            self.save_sale_to_excel(now, self.current_cart, total_amount)
            
            # Show invoice
            self.show_invoice(self.current_cart, total_books, total_amount, date_str, time_str)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate invoice: {str(e)}")
    
    def save_sale_to_excel(self, timestamp, cart, total):
        """Save sale to daily Excel file"""
        # Get date for filename
        date_str = timestamp.strftime("%d-%m-%Y")
        filename = f"Sales_Records/{date_str}.xlsx"
        
        # Prepare sale data
        sale_data = {
            'Date': timestamp.strftime("%d-%m-%Y"),
            'Time': timestamp.strftime("%I:%M %p"),
            'Book Title': [],
            'Class/Category': [],
            'SKU / Serial Number': [],
            'Unit Price (Rs)': [],
            'Total Bill (Rs)': []
        }
        
        # Add each book
        for i, book in enumerate(cart):
            sale_data['Book Title'].append(book['title'])
            sale_data['Class/Category'].append(f"Class {book['category']}")
            sale_data['SKU / Serial Number'].append(book['sku'])
            sale_data['Unit Price (Rs)'].append(f"Rs {book['price']:.2f}")
            
            # Only show total on first row
            if i == 0:
                sale_data['Total Bill (Rs)'].append(f"Rs {total:.2f}")
            else:
                sale_data['Total Bill (Rs)'].append('')
        
        # Check if file exists
        if os.path.exists(filename):
            # Append to existing file
            existing_df = pd.read_excel(filename)
            new_df = pd.DataFrame(sale_data)
            
            # Add separator row
            separator = pd.DataFrame([['---'] * len(sale_data)])
            separator.columns = existing_df.columns
            
            combined_df = pd.concat([existing_df, separator, new_df], ignore_index=True)
            combined_df.to_excel(filename, index=False)
        else:
            # Create new file
            df = pd.DataFrame(sale_data)
            df.to_excel(filename, index=False)
            
            # Format the Excel file
            wb = load_workbook(filename)
            ws = wb.active
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 22
            ws.column_dimensions['F'].width = 18
            
            wb.save(filename)
    
    def show_invoice(self, cart, total_books, total_amount, date, time):
        """Display the invoice"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="📄 Invoice Generated",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="New Sale",
            width=150,
            height=40,
            font=("Arial", 14),
            fg_color="#28a745",
            command=self.show_new_sale
        ).pack(side="right", padx=10)
        
        ctk.CTkButton(
            header,
            text="← Back to Menu",
            width=150,
            height=40,
            font=("Arial", 14),
            command=self.show_main_menu
        ).pack(side="right", padx=10)
        
        # Invoice container
        invoice_frame = ctk.CTkFrame(self.root)
        invoice_frame.pack(fill="both", expand=True, padx=50, pady=20)
        
        # Invoice header
        ctk.CTkLabel(
            invoice_frame,
            text="📚 SMART BOOK SHOP",
            font=("Arial", 24, "bold")
        ).pack(pady=10)
        
        ctk.CTkLabel(
            invoice_frame,
            text="SALES INVOICE",
            font=("Arial", 18, "bold")
        ).pack(pady=5)
        
        # Date and time
        info_frame = ctk.CTkFrame(invoice_frame)
        info_frame.pack(pady=15)
        
        ctk.CTkLabel(
            info_frame,
            text=f"Date: {date}  |  Time: {time}",
            font=("Arial", 14)
        ).pack()
        
        # Items table
        items_frame = ctk.CTkScrollableFrame(invoice_frame, height=300)
        items_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Header row
        header_row = ctk.CTkFrame(items_frame)
        header_row.pack(fill="x", pady=5)
        
        headers = ["Book Title", "Class", "SKU", "Price"]
        for header in headers:
            ctk.CTkLabel(
                header_row,
                text=header,
                font=("Arial", 14, "bold"),
                width=200
            ).pack(side="left", padx=10)
        
        # Items
        for book in cart:
            item_row = ctk.CTkFrame(items_frame)
            item_row.pack(fill="x", pady=2)
            
            values = [
                book['title'],
                f"Class {book['category']}",
                book['sku'],
                f"Rs {book['price']:.2f}"
            ]
            
            for value in values:
                ctk.CTkLabel(
                    item_row,
                    text=value,
                    font=("Arial", 12),
                    width=200
                ).pack(side="left", padx=10)
        
        # Summary
        summary_frame = ctk.CTkFrame(invoice_frame)
        summary_frame.pack(pady=20)
        
        ctk.CTkLabel(
            summary_frame,
            text=f"Total Books: {total_books}",
            font=("Arial", 16, "bold")
        ).pack(pady=5)
        
        ctk.CTkLabel(
            summary_frame,
            text=f"Total Amount: Rs {total_amount:.2f}",
            font=("Arial", 20, "bold"),
            text_color="#28a745"
        ).pack(pady=10)
        
        ctk.CTkLabel(
            invoice_frame,
            text="Thank you for your business!",
            font=("Arial", 14, "italic")
        ).pack(pady=10)
    
    # ============ SALES REPORTS ============
    
    def show_sales_reports(self):
        """Display sales reports interface"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="📊 Sales Reports",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_main_menu
        ).pack(side="right", padx=20)
        
        # Get all sales files
        sales_files = []
        if os.path.exists('Sales_Records'):
            sales_files = [f for f in os.listdir('Sales_Records') if f.endswith('.xlsx')]
            sales_files.sort(reverse=True)  # Most recent first
        
        if not sales_files:
            ctk.CTkLabel(
                self.root,
                text="No sales records found.",
                font=("Arial", 18)
            ).pack(expand=True)
            return
        
        # Reports list
        reports_frame = ctk.CTkFrame(self.root)
        reports_frame.pack(fill="both", expand=True, padx=50, pady=20)
        
        ctk.CTkLabel(
            reports_frame,
            text="Daily Sales Records:",
            font=("Arial", 18, "bold")
        ).pack(pady=20)
        
        list_frame = ctk.CTkScrollableFrame(reports_frame, height=500)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        for filename in sales_files:
            filepath = os.path.join('Sales_Records', filename)
            
            # Get file info
            try:
                df = pd.read_excel(filepath)
                num_sales = len([i for i in range(len(df)) if df.iloc[i, 0] != '---'])
                
                btn_frame = ctk.CTkFrame(list_frame)
                btn_frame.pack(fill="x", pady=5)
                
                date_display = filename.replace('.xlsx', '')
                info_text = f"📅 {date_display} - {num_sales} transaction(s)"
                
                ctk.CTkLabel(
                    btn_frame,
                    text=info_text,
                    font=("Arial", 14),
                    anchor="w"
                ).pack(side="left", fill="x", expand=True, padx=10)
                
                ctk.CTkButton(
                    btn_frame,
                    text="View Report",
                    width=120,
                    height=35,
                    command=lambda f=filepath: self.view_sales_report(f)
                ).pack(side="right", padx=10)
            except:
                pass
    
    def view_sales_report(self, filepath):
        """View a specific sales report"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        filename = os.path.basename(filepath).replace('.xlsx', '')
        ctk.CTkLabel(
            header,
            text=f"📊 Sales Report: {filename}",
            font=("Arial", 22, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_sales_reports
        ).pack(side="right", padx=20)
        
        try:
            # Load data
            df = pd.read_excel(filepath)
            
            # Display table
            table_frame = ctk.CTkFrame(self.root)
            table_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # Create Treeview
            columns = list(df.columns)
            tree = ttk.Treeview(
                table_frame,
                columns=columns,
                show='headings',
                height=25
            )
            
            # Headings
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=150)
            
            # Data
            for _, row in df.iterrows():
                values = [str(val) for val in row]
                tree.insert('', 'end', values=values)
            
            # Scrollbars
            vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            tree.grid(row=0, column=0, sticky='nsew')
            vsb.grid(row=0, column=1, sticky='ns')
            hsb.grid(row=1, column=0, sticky='ew')
            
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)
            
        except Exception as e:
            ctk.CTkLabel(
                self.root,
                text=f"Error loading report: {str(e)}",
                font=("Arial", 16)
            ).pack(expand=True)
    
    # ============ CHANGE PASSWORD ============
    
    def show_change_password(self):
        """Display change password interface"""
        self.clear_screen()
        
        # Header
        header = ctk.CTkFrame(self.root, height=70)
        header.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="🔐 Change Password",
            font=("Arial", 24, "bold")
        ).pack(side="left", padx=20)
        
        ctk.CTkButton(
            header,
            text="← Back",
            width=120,
            height=40,
            font=("Arial", 14),
            command=self.show_main_menu
        ).pack(side="right", padx=20)
        
        # Form
        form_frame = ctk.CTkFrame(self.root)
        form_frame.pack(expand=True, padx=100, pady=50)
        
        ctk.CTkLabel(
            form_frame,
            text="Current Password:",
            font=("Arial", 16)
        ).pack(pady=10)
        
        self.current_pass_entry = ctk.CTkEntry(
            form_frame,
            width=350,
            height=40,
            font=("Arial", 14),
            show="●"
        )
        self.current_pass_entry.pack(pady=10)
        
        ctk.CTkLabel(
            form_frame,
            text="New Password:",
            font=("Arial", 16)
        ).pack(pady=10)
        
        self.new_pass_entry = ctk.CTkEntry(
            form_frame,
            width=350,
            height=40,
            font=("Arial", 14),
            show="●"
        )
        self.new_pass_entry.pack(pady=10)
        
        ctk.CTkLabel(
            form_frame,
            text="Confirm New Password:",
            font=("Arial", 16)
        ).pack(pady=10)
        
        self.confirm_pass_entry = ctk.CTkEntry(
            form_frame,
            width=350,
            height=40,
            font=("Arial", 14),
            show="●"
        )
        self.confirm_pass_entry.pack(pady=10)
        
        ctk.CTkButton(
            form_frame,
            text="Change Password",
            width=250,
            height=50,
            font=("Arial", 16, "bold"),
            fg_color="#28a745",
            command=self.change_password
        ).pack(pady=30)
    
    def change_password(self):
        """Change the staff password"""
        current = self.current_pass_entry.get()
        new = self.new_pass_entry.get()
        confirm = self.confirm_pass_entry.get()
        
        if current != self.credentials['password']:
            messagebox.showerror("Error", "Current password is incorrect!")
            return
        
        if not new:
            messagebox.showerror("Error", "New password cannot be empty!")
            return
        
        if new != confirm:
            messagebox.showerror("Error", "Passwords do not match!")
            return
        
        self.credentials['password'] = new
        self.save_credentials()
        
        messagebox.showinfo("Success", "Password changed successfully!")
        self.show_main_menu()
    
    def run(self):
        """Start the application"""
        self.root.mainloop()


# ============ MAIN ENTRY POINT ============

if __name__ == "__main__":
    app = BookShopSystem()
    app.run()
